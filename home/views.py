from django.shortcuts import render , redirect
from django.contrib import messages
from home.filters import ExportFilter, ExportProductFilter
from django.contrib.auth.decorators import login_required
from django.db.models import Sum , Q, Count
from product.models import *
from home.models import *
from order.models import *
from home.forms import *
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from .forms import ContactForm ,SearchForm, EmailSignupForm
from django.template.loader import render_to_string, get_template
from django.core.paginator import Paginator , EmptyPage, PageNotAnInteger
from django.core.mail import send_mail
from django.conf import settings
from datetime import datetime , timedelta

import xlwt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import csv,io 
import json
import requests
# Create your views here.

MAILCHIMP_API_KEY = settings.MAILCHIMP_API_KEY
MAILCHIMP_DATA_CENTER = settings.MAILCHIMP_DATA_CENTER
MAILCHIMP_EMAIL_LIST_ID = settings.MAILCHIMP_EMAIL_LIST_ID

# api_url = f'https://{MAILCHIMP_DATA_CENTER}.api.mailchimp.com/3.0/'
api_url = f'https://{MAILCHIMP_DATA_CENTER}.api.mailchimp.com/3.0'
members_endpoint = f'{api_url}/lists/{MAILCHIMP_EMAIL_LIST_ID}/members'

def subscribe(email):
    data = {
        "email_address": email,
        "status": "subscribed"
    }
    r = requests.post(
        members_endpoint,
        auth=("", MAILCHIMP_API_KEY),
        data=json.dumps(data)
    )
    return r.status_code, r.json()


def email_list_signup(request):
    form = EmailSignupForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            email_signup_qs = Subscribe.objects.filter(email=form.instance.email)
            if email_signup_qs.exists():
                messages.info(request, "Email đã được đăng ký!")
            else:
                subscribe(form.instance.email)
                form.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def base(request):
    category = Category.objects.all()
    product_slider = Product.objects.all().order_by('-id')[:4]
    news_slider = News.objects.all().order_by('-id')[:4]
    product_laster = Product.objects.all().order_by('id')[:10]
    product_firster = Product.objects.all().order_by('-id')[:10]
    product_picker = Product.objects.all().order_by('?')[:10]
    cmt = Comment.objects.all().order_by('-id')[:10]
    form = EmailSignupForm()
    products = Product.objects.all()

    ob=News.objects.get(id=2)

    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    context = {
        'category' : category,
        'product_slider' : product_slider,
        'news_slider' : news_slider,
        'product_laster' : product_laster,
        'product_firster' : product_firster,
        'product_picker' : product_picker,
        'products' : products,
        'total' : total,
        'ob' : ob,
        'cmt' : cmt,
        'form' : form,
    }
    return render(request,'base.html',context)

def blank(request):
    category = Category.objects.all()
    product_slider = Product.objects.all().order_by('-id')[:4]
    news_slider = News.objects.all().order_by('-id')[:4]
    product_laster = Product.objects.all().order_by('-id')[:4]
    product_firster = Product.objects.all().order_by('id')[:4]
    product_picker = Product.objects.all().order_by('?')[:4]

    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    context = {
        'category' : category,
        'product_slider' : product_slider,
        'news_slider' : news_slider,
        'product_laster' : product_laster,
        'product_firster' : product_firster,
        'product_picker' : product_picker,
        'total' : total,
    }
    return render(request,'blank.html',context)

def password_reset_done(request):
    category = Category.objects.all()
    product_slider = Product.objects.all().order_by('-id')[:4]
    news_slider = News.objects.all().order_by('-id')[:4]
    product_laster = Product.objects.all().order_by('-id')[:4]
    product_firster = Product.objects.all().order_by('id')[:4]
    product_picker = Product.objects.all().order_by('?')[:4]
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    context = {
        'category' : category,
        'product_slider' : product_slider,
        'news_slider' : news_slider,
        'product_laster' : product_laster,
        'product_firster' : product_firster,
        'product_picker' : product_picker,
        'total' : total,
    }
    return render(request,'password_reset_done.html',context)

def password_reset(request):
    category = Category.objects.all()
    product_slider = Product.objects.all().order_by('-id')[:4]
    news_slider = News.objects.all().order_by('-id')[:4]
    product_laster = Product.objects.all().order_by('-id')[:4]
    product_firster = Product.objects.all().order_by('id')[:4]
    product_picker = Product.objects.all().order_by('?')[:4]
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    context = {
        'category' : category,
        'product_slider' : product_slider,
        'news_slider' : news_slider,
        'product_laster' : product_laster,
        'product_firster' : product_firster,
        'product_picker' : product_picker,
        'total' : total,
    }
    return render(request,'password_reset.html',context)

def password_reset_complete(request):
    category = Category.objects.all()
    product_slider = Product.objects.all().order_by('-id')[:4]
    news_slider = News.objects.all().order_by('-id')[:4]
    product_laster = Product.objects.all().order_by('-id')[:4]
    product_firster = Product.objects.all().order_by('id')[:4]
    product_picker = Product.objects.all().order_by('?')[:4]
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    context = {
        'category' : category,
        'product_slider' : product_slider,
        'news_slider' : news_slider,
        'product_laster' : product_laster,
        'product_firster' : product_firster,
        'product_picker' : product_picker,
        'total' : total,
    }
    return render(request,'password_reset_complete.html',context)



# def unsubscribe(request):
#     if request.method == "POST":
#         mail = request.POST['email']
#         form = SubscribeForm(request.POST or None)
#         if form.is_valid():
#             data = Subscribe()
#             data.email = form.cleaned_data['email']
#             if Subscribe.objects.filter(email=mail).exists():
#                 Subscribe.objects.filter(email=mail).delete()
#                 messages.success(request,"Đã hủy đăng ký!")
#                 data.save()
#                 return HttpResponseRedirect('/')
#             else:              
#                 messages.success(request,"Email chưa được đăng ký!")
#                 return HttpResponseRedirect('/')
#     else:
#         form = SubscribeForm()
#     context = {
#         'form' : form,
#     } 
#     return HttpResponseRedirect('/')

def contact(request):
    if request.method == "POST":
        na = request.POST['name']
        mess = request.POST['message']
        send_mail(na,mess,settings.EMAIL_HOST_USER,['fbworld.giang@gmail.com'],fail_silently=False)
        form = ContactForm(request.POST)
        if form.is_valid():
            data = ContactMessage()
            data.name = form.cleaned_data['name']
            data.email = form.cleaned_data['email']
            data.phone = form.cleaned_data['phone']
            data.message = form.cleaned_data['message']
            data.ip = request.META.get('REMOTE_ADDR')

            data.save()
        messages.success(request,"Gửi tin nhắn thành công! Cảm ơn bạn đã liên hệ với chúng tôi")
        return HttpResponseRedirect('/contact')
    else:
        form = ContactForm()
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    category = Category.objects.all()
    context = {
        'category' : category,
        'form' : form,
        'total' : total,

    }
    
    return render(request,'contact.html',context)

def category_products(request,id,slug):
    products = Product.objects.filter(category_id = id)
    c = products.count
    trademark = TradeMark.objects.all()
    paginator = Paginator(products, 10)
    pageNumber = request.GET.get('page')
    try:
        cmts = paginator.page(pageNumber)
    except PageNotAnInteger:
        cmts = paginator.page(1)
    except EmptyPage:
        cmts = paginator.page(paginator.num_pages)
    count = products.count()
    category = Category.objects.all()
    form = EmailSignupForm()
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    context = {
        'category' : category,
        'products' : products,
        'total' : total,
        'count' : count,
        'form' : form,
        'c' : c,
        'cmts' : cmts,
        'trademark' : trademark

    }
    return render(request,'category.html',context)


def trademark(request,id):
    products = Product.objects.filter(trademark_id = id)
    trademark = TradeMark.objects.all()
    c = products.count
    category = Category.objects.all()
    form = EmailSignupForm()
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    context = {
        'category' : category,
        'products' : products,
        'total' : total,
        'form' : form,
        'c' : c,
        'trademark' : trademark,

    }
    return render(request,'category.html',context)

def is_valid_queryparam(param):
    return param != '' and param is not None

def shop(request):
    products = Product.objects.all()
    category = Category.objects.all()
    trademark = TradeMark.objects.all()
    form = EmailSignupForm()
    product_laster = Product.objects.all().order_by('-id')[:7]
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    qs = Product.objects.all()
    title_contains_query = request.GET.get('title_contains')
    id_exact_query = request.GET.get('id_exact')
    title_or_author_query = request.GET.get('title_or_author')
    view_price_min = request.GET.get('view_price_min')
    view_price_max = request.GET.get('view_price_max')
    categorys = request.GET.get('categorys')

    if is_valid_queryparam(title_contains_query):
        qs = qs.filter(title__icontains = title_contains_query)

    elif is_valid_queryparam(id_exact_query):
        qs = qs.filter(id=id_exact_query)
    
    elif is_valid_queryparam(title_or_author_query):
        qs = qs.filter(Q(title__icontains=title_or_author_query)
                        | Q(author__name__icontains=title_or_author_query)
                        ).distinct()
    if is_valid_queryparam(view_price_min):
        qs = qs.filter(price__gte=view_price_min)

    if is_valid_queryparam(view_price_max):
        qs = qs.filter(price__lt=view_price_max)
    
    if is_valid_queryparam(categorys) and categorys != 'Chọn...':
        qs = qs.filter(trademark__name=categorys)
    c = qs.count
    paginator = Paginator(qs, 15)
    try:
        page = int(request.GET.get('page','1'))
    except:
        page = 1

    try:
        items = paginator.page(page)
    except(EmptyPage, InvalidPage):
        items = paginator.page(paginator.num_pages)

    context = {
        'category' : category,
        'products' : products,
        'product_laster' : product_laster,
        'total' : total,
        'trademark' : trademark,
        'queryset' : qs,
        'items' : items,
        'form' : form,
        'c' : c,
    }
    return render(request,'shop.html',context)


def shop_products(request,id,slug):
    products = Product.objects.all()
    category = Category.objects.all()
    trademark = TradeMark.objects.all()
    form = EmailSignupForm()
    product_laster = Product.objects.all().order_by('-id')[:7]
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    qs = Product.objects.all()
    title_contains_query = request.GET.get('title_contains')
    id_exact_query = request.GET.get('id_exact')
    title_or_author_query = request.GET.get('title_or_author')
    view_price_min = request.GET.get('view_price_min')
    view_price_max = request.GET.get('view_price_max')
    categorys = request.GET.get('categorys')

    if is_valid_queryparam(title_contains_query):
        qs = qs.filter(title__icontains = title_contains_query)

    elif is_valid_queryparam(id_exact_query):
        qs = qs.filter(id=id_exact_query)
    
    elif is_valid_queryparam(title_or_author_query):
        qs = qs.filter(Q(title__icontains=title_or_author_query)
                        | Q(author__name__icontains=title_or_author_query)
                        ).distinct()
    if is_valid_queryparam(view_price_min):
        qs = qs.filter(price__gte=view_price_min)

    if is_valid_queryparam(view_price_max):
        qs = qs.filter(price__lt=view_price_max)
    
    if is_valid_queryparam(categorys) and categorys != 'Chọn...':
        qs = qs.filter(trademark__name=categorys)
    c = qs.count
    paginator = Paginator(qs, 15)
    try:
        page = int(request.GET.get('page','1'))
    except:
        page = 1

    try:
        items = paginator.page(page)
    except(EmptyPage, InvalidPage):
        items = paginator.page(paginator.num_pages)

    context = {
        'category' : category,
        'products' : products,
        'product_laster' : product_laster,
        'total' : total,
        'trademark' : trademark,
        'queryset' : qs,
        'items' : items,
        'form' : form,
        'c' : c,
    }
    return render(request,'shop.html',context)



def search(request):
    if request.method == 'POST': # check post
        form = SearchForm(request.POST)
        if form.is_valid():
            q = form.cleaned_data['query'] # get form input data
            products=Product.objects.filter(title__icontains=q)
            c = products.count
            trademark = TradeMark.objects.all()
            paginator = Paginator(products, 10)
            pageNumber = request.GET.get('page')
            try:
                cmts = paginator.page(pageNumber)
            except PageNotAnInteger:
                cmts = paginator.page(1)
            except EmptyPage:
                cmts = paginator.page(paginator.num_pages)
            category = Category.objects.all()
            if products:
                category = Category.objects.all()
                context = {
                            'products': products,
                            'q':q,
                            'category': category,
                            'c': c,
                            'trademark': trademark,
                            'cmts': cmts,
                            }
                return render(request, 'products_search.html', context)
            else:
                messages.error(request,"Không tìm thấy")
                return HttpResponseRedirect('/')                    
    return HttpResponseRedirect('/')

# def search_auto(request):
#     if request.is_ajax():
#         q = request.GET.get('term', '')
#         products = Product.objects.filter(title__icontains=q)

#         results = []
#         for rs in products:
#             product_json = {}
#             product_json = rs.title +" > " + rs.category.title
#             results.append(product_json)
#         data = json.dumps(results)
#     else:
#         data = 'fail'
#     mimetype = 'application/json'
#     return HttpResponse(data, mimetype)


def product_detail(request,id,slug):  
    category = Category.objects.all()
    product = Product.objects.get(pk=id)
    images = Images.objects.filter(product_id=id)
    comments = Comment.objects.filter(product_id=id,status='True')
    paginator = Paginator(comments, 3)
    pageNumber = request.GET.get('page')
    try:
        cmts = paginator.page(pageNumber)
    except PageNotAnInteger:
        cmts = paginator.page(1)
    except EmptyPage:
        cmts = paginator.page(paginator.num_pages)

    product_laster = Product.objects.all().order_by('-id')[:7]
    product_firster = Product.objects.all().order_by('id')[:4]

    query = request.GET.get('q')
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity

    context = {
        'category' : category,
        'product' : product,
        'images' : images,
        'product_laster' : product_laster,
        'product_firster' : product_firster,
        'total' : total,
        'comments': comments,
        'cmts': cmts,
    }
    if product.variant != "None":
        if request.method == 'POST': #if we select color
            variant_id = request.POST.get('variantid')
            variant = Variants.objects.get(id = variant_id)
            colors = Variants.objects.filter(product_id = id,size_id = variant.size_id)
            print(colors)
            sizes = Variants.objects.raw('SELECT * FROM product_variants WHERE product_id = %s GROUP BY size_id',[id])
            query += variant.title +'Size:' + str(variant.size) + ' Color:' + str(variant.color)
        
        else:
            variants = Variants.objects.filter(product_id=id)
            colors = Variants.objects.filter(product_id=id,size_id=variants[0].size_id )
            sizes = Variants.objects.raw('SELECT * FROM  product_variants  WHERE product_id=%s GROUP BY size_id',[id])
            variant =Variants.objects.get(id=variants[0].id)
        context.update({
            'sizes' : sizes,
            'colors' : colors,
            'variant' : variant,
            'query' : query,
        })

    return render(request,'product_detail.html',context)



def ajaxcolor(request):
    data = {}
    if request.POST.get('action') == 'POST':
        size_id = request.POST.get('size')
        productid = request.POST.get('productid')
        colors = Variants.objects.filter(product_id = productid, size_id = size_id)
        context = {
            'size_id' : size_id,
            'productid' : productid,
            'colors' : colors,
        }
        data = {'rendered_table' : render_to_string('color_list.html',context = context)}
        return JsonResponse(data)
    return JsonResponse(data)



def get_total_per_month_value():
    result= {}
    db_result = Order.objects.values('total','create_at')
    for i in db_result:
        year = str(i.get('create_at').strftime("%Y"))
        # if year == 2020:
        month = str(i.get('create_at').strftime("%m"))
        if month in result.keys():
            result[month] = result[month] + i.get('total')
        else:
            result[month] = i.get('total')
        # else:
        #     messages.success(request, "Không tìm thấy kết quả phù hợp")
        #     return HttpResponseRedirect('/chart')
        
    return result


@login_required(login_url='/login')
def chart(request):
    category = Category.objects.all()
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    a = get_total_per_month_value()
    print (a)
    #month = Order.objects.raw("SELECT strftime('%m',order_order.create_at)AS thang FROM order_order WHERE strftime('%Y',order_order.create_at)='2020'")
    #money = Order.objects.raw("SELECT SUM (order_order.total)AS sotien FROM order_order WHERE strftime('%Y',order_order.create_at)='2020' GROUP BY strftime('%m',order_order.create_at)")
    # money = Order.objects.filter(create_at__month= 6).aggregate(Sum('total'))

    # invoices = Order.objects.all()
    # months = invoices.datetimes("create_at", kind="month")
    # for month in months:
    #     month_invs = invoices.filter(create_at__month=month.month)
    #     month_total = month_invs.aggregate(total=Sum("total"))
    #     print(f"Month: {month}, Total: {month_total}")


    
    # b = int('0' + '06')
    # print (b)
    # listthang = []
    # for k in range(1,13):
    #     lt ={'k':k}
    #     listthang.append(lt)
    # thang = [x['k'] for x in listthang]

    # tongtien = 0
    # listtien = []
    # for rs in thang:
    #     if rs != None:
    #         tien = Order.objects.filter(create_at__month= rs).aggregate(Sum('total'))
    #         for h in tien:
    #             ti = {'h':h}
            # if tien == None:
            #     tien = 0
            #     print("1")
            # print(tien)
            
    #         listtien.append(ti)
    #     t = [x['h'] for x in listtien]
    # print(t)
    # print(tien)
    # for k in tien:
    #     # money = Order.objects.filter(create_at__month= 6).aggregate(Sum('total'))
    #     # listtien.append(money)
    #     # tongtien += k.total
    #     print(k.total)
    #     tongtien += k.total
    #     # t = {
    #     #     'SoTien':k.total
    #     # }
    #     # listtien.append(t)
    # listtien.append(tongtien)
    # # tongtien = [x['SoTien'] for x in listtien]
    # # 0,0,0,0,918000
    # print(tongtien)
    context = {
        'category' : category,
        'shopcart': shopcart,
        'total': total,
        # 'thang': thang,
        'a': a,
        # 'b': b,
        # 'month_total': month_total,
        # 'tongtien': tongtien,
        # 'output': column2D.render(),
    }
    return render(request,'chart.html',context)
 
# cart

@login_required(login_url='/login')
def addtoshopcart(request, id):
    url = request.META.get('HTTP_REFERER')  # get last url
    current_user = request.user  # Access User Session information
    variantid = request.POST.get('variantid')  # from variant add to cart
    checkinvariant = ShopCart.objects.filter(variant_id=variantid)  # Check product in shopcart

    control = 0
    # if checkinproduct:
    #     control = 1
    if checkinvariant:
        control = 1
    else:
        control = 0

    if request.method == 'POST':  # if there is a post
        form = ShopCartForm(request.POST)
        if form.is_valid():
            if control==1: # Update  shopcart
                data = ShopCart.objects.get(variant_id=variantid)
                data.quantity += form.cleaned_data['quantity']
                data.save()  # save data
            else : # Inser to Shopcart
                data = ShopCart()
                data.user_id = current_user.id
                data.product_id =id
                data.variant_id = variantid
                data.quantity = form.cleaned_data['quantity']
                data.save()
        messages.success(request, "Đã thêm sản phẩm vào giỏ hàng")
        return HttpResponseRedirect(url)

    else: # if there is no post
        if control == 1:  # Update  shopcart
            data = ShopCart.objects.get(variant_id=id) 
            data.quantity += 1
            data.save()  #
        else:  #  Inser to Shopcart
            data = ShopCart()  
            data.user_id = current_user.id
            data.product_id = id
            data.quantity = 1
            data.variant_id =None
            data.save()  #
        messages.success(request, "Đã thêm sản phẩm vào giỏ hàng")
        return HttpResponseRedirect(url)

@login_required(login_url='/login')
def deletefromcart(request, id):
    # url = request.META.get('HTTP_REFERER')
    ShopCart.objects.filter(id=id ).delete()
    messages.success(request,"Đã xóa sản phẩm trong giỏ")
    return HttpResponseRedirect('/shopcart')


@login_required(login_url='/login')
def delete_single_item_from_cart(request,id):
    url = request.META.get('HTTP_REFERER')
    current_user = request.user
    checkproduct = ShopCart.objects.filter(variant_id = id)
    data = ShopCart.objects.get(variant_id=id)
    data.quantity -= 1
    data.save()
    messages.success(request, "Đã cập nhật giỏ hàng")
    return HttpResponseRedirect(url)

@login_required(login_url='/login')
def add_single_item_from_cart(request,id):
    url = request.META.get('HTTP_REFERER')
    current_user = request.user
    variantid = request.POST.get('variantid')  
    checkproduct = ShopCart.objects.filter(variant_id = id)
    data = ShopCart.objects.get(variant_id=id)
    data.quantity += 1
    data.save()
    messages.success(request, "Đã cập nhật giỏ hàng")
    return HttpResponseRedirect(url)

@login_required(login_url='/login')
def report(request):
    product = Product.objects.all()
    order_list = Order.objects.all()
    order_filter = ExportFilter(request.GET, queryset=order_list)

    
    context ={
        'product': product,
        'filter': order_filter,

    }
    return render(request, 'report.html', context)


def export_users_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="DoanhThu.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Doanh thu')


    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Mã', 'Tên tài khoản', 'Họ và đệm', 'Tên', 'Địa chỉ', 'Quận/Huyện', 'Tỉnh/Thành phố', 'Tổng tiền', 'Ghi chú', 'Trạng thái']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style) 

    font_style = xlwt.XFStyle()

    rows = Order.objects.filter(status='Hoàn thành').values_list('code', 'user', 'first_name', 'last_name', 'address', 'district', 'city', 'total','note','status')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)

    return response


@login_required(login_url='/login')
def report_products(request):
    product = Product.objects.all()
    variant_list = Variants.objects.all()
    product_filter = ExportProductFilter(request.GET, queryset=variant_list)
    context ={
        'product': product,
        'filter': product_filter
    }
    return render(request, 'report_products.html', context)

@login_required(login_url='/login')
def chart_2(request):
    category = Category.objects.all()
    current_user = request.user
    shopcart = ShopCart.objects.filter(user_id = current_user.id)
    total = 0
    for rs in shopcart:
        if rs.product.variant == 'None':
            total += rs.product.price * rs.quantity
        else:
            total += rs.variant.price * rs.quantity
    a = get_total_per_month_value()
    context = {
        'category' : category,
        'shopcart': shopcart,
        'total': total,
        'a': a,

    }
    return render(request,'chart_2.html',context)